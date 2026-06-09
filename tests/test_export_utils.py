"""
tests/test_export_utils.py
──────────────────────────
Tests for export_utils:
  • export_chat_markdown structure and content
  • export_results_to_xlsx sheet count and validity
"""

from __future__ import annotations

import io

import openpyxl
import pandas as pd
import pytest

from export_utils import export_chat_markdown, export_results_to_xlsx


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def simple_history() -> list[dict]:
    return [
        {"role": "user",      "content": "Show top 5 rows", "extra": {}},
        {"role": "assistant", "content": "Here is the table:",
         "extra": {"dataframe": pd.DataFrame({"a": [1, 2], "b": [3, 4]}), "code": "result = df.head()"}},
        {"role": "user",      "content": "What is the average?", "extra": {}},
        {"role": "assistant", "content": "**Result:** `3.5`",
         "extra": {"code": "result = df['a'].mean()"}},
    ]


@pytest.fixture
def empty_history() -> list[dict]:
    return []


# ── export_chat_markdown ──────────────────────────────────────────────────────

def test_markdown_contains_header(simple_history):
    md = export_chat_markdown(simple_history, "test.xlsx")
    assert "# Lumina transcript" in md


def test_markdown_contains_file_name(simple_history):
    md = export_chat_markdown(simple_history, "my_data.csv")
    assert "my_data.csv" in md


def test_markdown_contains_user_turns(simple_history):
    md = export_chat_markdown(simple_history, "f.csv")
    assert "Show top 5 rows" in md
    assert "What is the average?" in md


def test_markdown_contains_assistant_turns(simple_history):
    md = export_chat_markdown(simple_history, "f.csv")
    assert "Here is the table:" in md
    assert "Result:" in md


def test_markdown_contains_code_blocks(simple_history):
    md = export_chat_markdown(simple_history, "f.csv")
    assert "```python" in md
    assert "result = df.head()" in md


def test_markdown_empty_history_no_crash(empty_history):
    md = export_chat_markdown(empty_history, "f.csv")
    assert "Lumina transcript" in md


def test_markdown_ends_with_newline(simple_history):
    md = export_chat_markdown(simple_history, "f.csv")
    assert md.endswith("\n")


# ── export_results_to_xlsx ────────────────────────────────────────────────────

def test_xlsx_returns_bytes(simple_history):
    result = export_results_to_xlsx(simple_history)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_xlsx_is_valid_workbook(simple_history):
    xlsx_bytes = export_results_to_xlsx(simple_history)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb is not None


def test_xlsx_sheet_count_matches_dataframes(simple_history):
    """simple_history has 1 DataFrame result → 1 data sheet."""
    xlsx_bytes = export_results_to_xlsx(simple_history)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    sheet_names = [n for n in wb.sheetnames if n.startswith("Result_")]
    assert len(sheet_names) == 1


def test_xlsx_result_sheet_has_data(simple_history):
    xlsx_bytes = export_results_to_xlsx(simple_history)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Result_1"]
    # Header row + 2 data rows = at least 3 rows
    assert ws.max_row >= 3


def test_xlsx_empty_history_returns_placeholder(empty_history):
    xlsx_bytes = export_results_to_xlsx(empty_history)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    # Should have at least one sheet (the Info/placeholder sheet)
    assert len(wb.sheetnames) >= 1


def test_xlsx_multiple_dataframe_results():
    """Two DataFrame results → two Result sheets."""
    history = [
        {"role": "assistant", "content": "r1",
         "extra": {"dataframe": pd.DataFrame({"x": [1, 2]})}},
        {"role": "assistant", "content": "r2",
         "extra": {"dataframe": pd.DataFrame({"y": [3, 4]})}},
    ]
    xlsx_bytes = export_results_to_xlsx(history)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    sheet_names = [n for n in wb.sheetnames if n.startswith("Result_")]
    assert len(sheet_names) == 2
